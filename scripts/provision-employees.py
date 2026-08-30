#!/usr/bin/env python3
"""Provision LiteLLM users + virtual keys from config/employees.yaml.

Writes email-ready instruction files to secret/outbox/ (gitignored).
Rotates CHAT/AGENT keys by default so plaintext can be mailed once.

Usage:
  python3 scripts/provision-employees.py
  python3 scripts/provision-employees.py --no-rotate
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import secrets
import string
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import date
from pathlib import Path

try:
    import yaml
except ImportError:
    sys.stderr.write("Install pyyaml: pip install pyyaml\n")
    sys.exit(1)


def load_env(root: Path) -> None:
    env = root / ".env"
    if not env.exists():
        return
    for line in env.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        os.environ.setdefault(k, v)


def api(gateway: str, master: str, method: str, path: str, body: dict | None = None) -> dict:
    data = None if body is None else json.dumps(body).encode()
    req = urllib.request.Request(
        f"{gateway}{path}",
        data=data,
        method=method,
        headers={
            "Authorization": f"Bearer {master}",
            "Content-Type": "application/json",
            "User-Agent": "Mozilla/5.0 ImStemGateway/1.0",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            raw = resp.read().decode()
            return json.loads(raw) if raw else {}
    except urllib.error.HTTPError as e:
        err = e.read().decode()
        try:
            parsed = json.loads(err)
        except json.JSONDecodeError:
            parsed = {"raw": err}
        return {"error": e.code, "body": parsed}


def password(length: int = 14) -> str:
    alphabet = string.ascii_letters + string.digits
    return "".join(secrets.choice(alphabet) for _ in range(length))


def list_key_aliases(gateway: str, master: str) -> dict[str, dict]:
    found: dict[str, dict] = {}
    page = 1
    while page <= 20:
        result = api(
            gateway,
            master,
            "GET",
            f"/key/list?return_full_object=true&page={page}&size=100",
        )
        keys = result.get("keys") or result.get("data") or []
        if isinstance(keys, dict):
            keys = keys.get("keys") or []
        if not keys:
            break
        for item in keys:
            if not isinstance(item, dict):
                continue
            alias = item.get("key_alias")
            if alias:
                found[str(alias)] = item
        total_pages = int(result.get("total_pages") or 1)
        if page >= total_pages:
            break
        page += 1
    return found


def ensure_user(gateway: str, master: str, emp: dict, duration: str) -> None:
    emp_id = emp["id"]
    chat_b = float(emp.get("chat_budget_usd") or 40)
    agent_b = float(emp.get("agent_budget_usd") or 80)
    payload = {
        "user_id": emp_id,
        "user_alias": emp["name"],
        "user_email": emp.get("email") or f"{emp_id.lower()}@imstem.local",
        "max_budget": chat_b + agent_b,
        "budget_duration": duration,
        "metadata": {
            "department": emp.get("department"),
            "employee_id": emp_id,
            "title": emp.get("title"),
        },
    }
    created = api(gateway, master, "POST", "/user/new", payload)
    if created.get("error"):
        updated = api(gateway, master, "POST", "/user/update", payload)
        if updated.get("error"):
            print(f"{emp_id}: user update error {updated}")
        else:
            print(f"{emp_id}: user updated")
    else:
        print(f"{emp_id}: user created")


def issue_key(
    gateway: str,
    master: str,
    emp: dict,
    kind: str,
    models: list[str],
    budget: float,
    duration: str,
    existing: dict[str, dict],
    rotate: bool,
) -> str:
    emp_id = emp["id"]
    alias = f"{emp_id}-{kind}"
    if alias in existing and not rotate:
        print(f"{alias}: kept (no plaintext; pass nothing to rotate)")
        return ""
    if alias in existing and rotate:
        deleted = api(gateway, master, "POST", "/key/delete", {"key_aliases": [alias]})
        if deleted.get("error"):
            print(f"{alias}: delete error {deleted}")
    body = {
        "user_id": emp_id,
        "key_alias": alias,
        "models": models,
        "max_budget": budget,
        "budget_duration": duration,
        "metadata": {
            "kind": kind.lower(),
            "employee_id": emp_id,
            "department": emp.get("department"),
        },
    }
    result = api(gateway, master, "POST", "/key/generate", body)
    key = result.get("key") or ""
    if result.get("error") or not key:
        print(f"{alias}: generate error {result}")
        return ""
    print(f"{alias}: issued")
    return key


def instruction(
    emp: dict,
    chat_key: str,
    agent_key: str,
    webui_password: str,
    public: dict,
    models: list[str],
) -> str:
    emp_id = emp["id"]
    email = emp.get("email") or ""
    chat_url = public.get("chat_url", "https://chat.imstem.org")
    api_base = public.get("api_base", "https://llm.imstem.org/v1")
    anthropic = public.get("anthropic_base", "https://llm.imstem.org")
    guide = public.get("guide_url", "https://llm.imstem.org/guide")
    landing = public.get("landing_url", "https://llm.imstem.org")
    model_list = ", ".join(models)
    warn = ""
    if emp.get("email_inferred"):
        warn = (
            "\n> 注意：表格里没有公司邮箱，登录邮箱按姓名拼音推断为 "
            f"`{email}`。发信前请核对。\n"
        )
    return f"""# ImStem 大模型账号 — {emp['name']}（{emp_id}）

发给：{email or '(无邮箱)'}
{warn}
入口：{landing}

## 网页对话（Open WebUI）

- 地址：{chat_url}
- 登录邮箱：`{email}`
- 密码：`{webui_password}`
- 显示名：{emp['name']}

登录后在左上角选择模型。当前可选：

`{model_list}`

## 个人 API（LiteLLM 虚拟密钥）

这不是阿里云或小米的官方 Key。只能打 `{anthropic}`。

| 用途 | 别名 | 密钥 |
|---|---|---|
| 浏览器对话记账 | `{emp_id}-CHAT` | `{chat_key or '（未轮换，见管理员）'}` |
| Codex / Claude Code / Python | `{emp_id}-AGENT` | `{agent_key or '（未轮换，见管理员）'}` |

详细说明：{guide}

### Codex / OpenCode / Python

```bash
export OPENAI_BASE_URL={api_base}
export OPENAI_API_KEY={agent_key or 'sk-你的AGENT密钥'}
```

推荐模型：`qwen3.8-flash`（快）、`qwen3.8-max`（难）、`kimi-k3`（代码 / Agent）。

Python：

```python
from openai import OpenAI
client = OpenAI(base_url="{api_base}", api_key="{agent_key or 'sk-你的AGENT密钥'}")
print(client.chat.completions.create(
    model="qwen3.8-flash",
    messages=[{{"role": "user", "content": "你好"}}],
).choices[0].message.content)
```

### Claude Code

不要在 Base URL 后面加 `/v1`。

```bash
export ANTHROPIC_BASE_URL={anthropic}
export ANTHROPIC_API_KEY={agent_key or 'sk-你的AGENT密钥'}
export ANTHROPIC_MODEL=kimi-k3
```

如果工具仍发送 `gpt-4o` / `gpt-5` / `claude-sonnet-4-5`，网关会转到 `kimi-k3`，费用记在你名下。

## 注意

- 不要把这把密钥转给同事
- 不要上传可识别患者信息、密码、未脱敏病历
- 月度额度用完后密钥会暂时失败，找管理员加额度
"""


def update_xlsx(root: Path, roster: list[dict]) -> None:
    path = root / "secret" / "users.xlsx"
    if not path.exists():
        return
    try:
        import openpyxl
    except ImportError:
        print("openpyxl missing; skip users.xlsx update")
        return
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {str(h): i for i, h in enumerate(headers) if h}
    by_name = {r["name"]: r for r in roster}
    today = date.today().isoformat()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        name = row[idx.get("姓名", 1)].value
        if not name or name not in by_name:
            continue
        rec = by_name[name]
        if "Employee ID" in idx:
            row[idx["Employee ID"]].value = rec["id"]
        if "OpenWebUI账号" in idx:
            row[idx["OpenWebUI账号"]].value = rec["email"]
        if "启用日期" in idx and not row[idx["启用日期"]].value:
            row[idx["启用日期"]].value = today
        if rec.get("email_inferred") and "管理员备注" in idx:
            row[idx["管理员备注"]].value = (
                (row[idx["管理员备注"]].value or "")
                + " 邮箱按拼音推断，发信前核对"
            ).strip()
    wb.save(path)
    print(f"updated {path}")


def update_openwebui_shared_key(gateway: str, master: str, models: list[str]) -> None:
    existing = list_key_aliases(gateway, master)
    item = existing.get("OPENWEBUI-CHAT")
    if not item:
        print("OPENWEBUI-CHAT: not found")
        return
    token = item.get("token") or item.get("key")
    if not token:
        print("OPENWEBUI-CHAT: no token in list")
        return
    result = api(
        gateway,
        master,
        "POST",
        "/key/update",
        {"key": token, "models": models},
    )
    print("OPENWEBUI-CHAT models", "ok" if not result.get("error") else result)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--no-rotate", action="store_true")
    args = parser.parse_args()

    root = Path(__file__).resolve().parents[1]
    load_env(root)
    gateway = (os.environ.get("PUBLIC_GATEWAY_URL") or "https://llm.imstem.org").rstrip("/")
    master = os.environ.get("LITELLM_MASTER_KEY")
    if not master:
        sys.stderr.write("LITELLM_MASTER_KEY missing\n")
        return 1

    cfg = yaml.safe_load((root / "config" / "employees.yaml").read_text())
    public = cfg.get("public") or {}
    duration = cfg.get("budget_duration", "30d")
    chat_models = list(cfg.get("chat_models") or [])
    agent_models = json.loads((root / "config" / "agent-models.json").read_text())["agent_models"]

    for name, dept in (cfg.get("departments") or {}).items():
        api(
            gateway,
            master,
            "POST",
            "/team/new",
            {
                "team_alias": dept.get("team_alias") or name,
                "models": chat_models,
                "metadata": {"department": name},
            },
        )

    existing = list_key_aliases(gateway, master)
    update_openwebui_shared_key(gateway, master, chat_models)

    outdir = root / "secret" / "outbox"
    outdir.mkdir(parents=True, exist_ok=True)
    roster: list[dict] = []
    rotate = not args.no_rotate

    for emp in cfg.get("employees") or []:
        ensure_user(gateway, master, emp, duration)
        chat_key = issue_key(
            gateway,
            master,
            emp,
            "CHAT",
            chat_models,
            float(emp.get("chat_budget_usd") or 40),
            duration,
            existing,
            rotate,
        )
        agent_key = issue_key(
            gateway,
            master,
            emp,
            "AGENT",
            agent_models,
            float(emp.get("agent_budget_usd") or 80),
            duration,
            existing,
            rotate,
        )
        webui_password = password()
        text = instruction(emp, chat_key, agent_key, webui_password, public, chat_models)
        (outdir / f"{emp['id']}.md").write_text(text, encoding="utf-8")
        rec = {
            "id": emp["id"],
            "name": emp["name"],
            "email": emp.get("email") or "",
            "email_inferred": bool(emp.get("email_inferred")),
            "department": emp.get("department"),
            "openwebui_login": emp.get("email") or "",
            "openwebui_password": webui_password,
            "chat_key": chat_key,
            "agent_key": agent_key,
        }
        roster.append(rec)

    roster_path = outdir / "roster.csv"
    with roster_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(
            fh,
            fieldnames=[
                "id",
                "name",
                "email",
                "email_inferred",
                "department",
                "openwebui_login",
                "openwebui_password",
                "chat_key",
                "agent_key",
            ],
        )
        writer.writeheader()
        writer.writerows(roster)
    (outdir / "README.md").write_text(
        "These files are secrets. Email one `{ID}.md` per person. Do not commit.\n",
        encoding="utf-8",
    )
    update_xlsx(root, roster)
    print(f"wrote {outdir}")
    print("Next: create matching Open WebUI users (see scripts/create-openwebui-users.py).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
